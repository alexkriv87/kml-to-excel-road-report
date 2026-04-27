# excel_exporter.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import re
from pathlib import Path
from constants import ROAD_PRIORITY, ROAD_COLOR
from logger_config import setup_logging

logger = setup_logging()

# Константы для расчета высоты строк
CHARS_PER_LINE = 65      # реально помещается в строку
BASE_HEIGHT = 60         # высота для 3 строк
EXTRA_HEIGHT_PER_LINE = 20  # добавляем на каждую доп. строку


def export_to_excel(gdf, output_file, template_path, include_geometry=False):
    """
    Экспортирует данные в Excel на основе шаблона.

    Параметры:
        gdf: GeoDataFrame с данными
        output_file: путь для сохранения результата
        template_path: путь к файлу-шаблону
        include_geometry: включать ли колонку geometry (для отладки)
    """
    # Преобразуем пути в Path объекты
    output_file = Path(output_file)
    template_path = Path(template_path)

    # 1. Открываем шаблон
    workbook = openpyxl.load_workbook(template_path)
    worksheet = workbook['ИД']

    # 2. Подготовка данных
    excel_df = gdf.copy()
    excel_df['Код'] = excel_df['Принадлежность'].map(ROAD_PRIORITY).fillna(0)

    # 3. Стили
    font_tnr = Font(name='Times New Roman', size=11)

    # Для центрированных колонок (D, F, G, I, K, L, M, N, O, P, S, T, U, V, W)
    align_center = Alignment(horizontal='center', vertical='center')

    # Для колонок с длинным текстом и переносом (E, Q, R)
    align_left_wrap = Alignment(
        horizontal='left', vertical='center', wrap_text=True)

    # Для колонки H (центр + перенос)
    align_center_wrap = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 4. Заполняем данные (начиная со строки 2)
    for row_num, row in excel_df.iterrows():
        excel_row = row_num + 2  # +2 потому что строка 1 - заголовки

        # D - Номер (центр)
        cell = worksheet[f'D{excel_row}']
        cell.value = row['Номер']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # E - Название (влево, с переносом)
        cell = worksheet[f'E{excel_row}']
        cell.value = row['Название']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_left_wrap         # влево + перенос
        cell.border = thin_border                # тонкие границы

        # F - категория (центр)
        cell = worksheet[f'F{excel_row}']
        cell.value = row['категория']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # G - покрытие (центр)
        cell = worksheet[f'G{excel_row}']
        cell.value = row['покрытие']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # H - принадлежность из description (центр, с переносом)
        cell = worksheet[f'H{excel_row}']
        cell.value = row['принадлежность']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center_wrap       # центр + перенос
        cell.border = thin_border                # тонкие границы

        # I - осевая нагрузка (центр)
        cell = worksheet[f'I{excel_row}']
        value = row['осевая_нагрузка']
        if pd.notna(value):
            cell.value = value
            if value == int(value):
                cell.number_format = '0'         # целое число
            else:
                cell.number_format = '0.0'       # один знак
        else:
            cell.value = ''
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # J - Используемый участок (формула)
        cell = worksheet[f'J{excel_row}']
        cell.value = f'="от "&Q{excel_row}&" до "&R{excel_row}'
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center_wrap       # центр + перенос
        cell.border = thin_border                # тонкие границы

        # K - ширина (центр)
        cell = worksheet[f'K{excel_row}']
        cell.value = row['ширина']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # L - Протяженность (центр)
        cell = worksheet[f'L{excel_row}']
        cell.value = row['Протяженность']
        cell.number_format = '0.00'              # два знака
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # M - Фото (пусто, центр)
        cell = worksheet[f'M{excel_row}']
        cell.value = ''
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # N - Принадлежность (тип дороги, центр)
        cell = worksheet[f'N{excel_row}']
        cell.value = row['Принадлежность']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # O - Код (центр)
        cell = worksheet[f'O{excel_row}']
        cell.value = row['Код']
        cell.number_format = '0'                 # целое число
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # P - гиперссылка на лист дороги с проверкой AE
        cell = worksheet[f'P{excel_row}']
        cell.value = f'=HYPERLINK("#\'"&D{excel_row}&"\'!A1", IF(AND(AE{excel_row}<>"", AE{excel_row}=0), " Маршрут заполнен", " Маршрут не заполнен"))'
        cell.font = Font(name='Times New Roman', size=11,
                         underline='single')  # шрифт Times New Roman 11
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border                # тонкие границы

        # Q - Откуда (влево, с переносом)
        cell = worksheet[f'Q{excel_row}']
        cell.value = row['Откуда']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_left_wrap         # влево + перенос
        cell.border = thin_border                # тонкие границы

        # R - Куда (влево, с переносом)
        cell = worksheet[f'R{excel_row}']
        cell.value = row['Куда']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_left_wrap         # влево + перенос
        cell.border = thin_border                # тонкие границы

        # S - пустая (центр)
        cell = worksheet[f'S{excel_row}']
        cell.value = ''
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # T - Начало (широта) (центр)
        cell = worksheet[f'T{excel_row}']
        cell.value = row['Начало (широта)']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # U - Начало (долгота) (центр)
        cell = worksheet[f'U{excel_row}']
        cell.value = row['Начало (долгота)']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # V - Конец (широта) (центр)
        cell = worksheet[f'V{excel_row}']
        cell.value = row['Конец (широта)']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # W - Конец (долгота) (центр)
        cell = worksheet[f'W{excel_row}']
        cell.value = row['Конец (долгота)']
        cell.font = font_tnr                    # шрифт Times New Roman 11
        cell.alignment = align_center            # выравнивание по центру
        cell.border = thin_border                # тонкие границы

        # X, Y, Z - не заполняем, они останутся из шаблона

        # Geometry для отладки (если нужно)
        if include_geometry:
            cell = worksheet[f'AA{excel_row}']
            cell.value = str(row.geometry)
            cell.font = font_tnr                # шрифт Times New Roman 11
            cell.alignment = align_center        # выравнивание по центру
            cell.border = thin_border            # тонкие границы

        # Расчет высоты строки по содержимому J
        j_text = f"от {row['Откуда']} до {row['Куда']}"
        j_length = len(j_text)

        lines = (j_length + CHARS_PER_LINE -
                 1) // CHARS_PER_LINE  # округление вверх

        if lines <= 3:
            row_height = BASE_HEIGHT
        else:
            extra_lines = lines - 3
            row_height = BASE_HEIGHT + extra_lines * EXTRA_HEIGHT_PER_LINE

        worksheet.row_dimensions[excel_row].height = row_height

        # AD - вспомогательный столбец (протяженность дороги из листа дороги ячейка С11)
        cell = worksheet[f'AD{excel_row}']
        cell.value = f'=IFERROR(INDIRECT("\'"&D{excel_row}&"\'!K3"), "нет данных")'

        # AE - Расчет разницы протяженности дороги (заполненная и с саски)
        cell = worksheet[f'AE{excel_row}']
        cell.value = f'=IF(AND(L{excel_row}<>"", AD{excel_row}<>""), L{excel_row}-AD{excel_row}, "")'

    # 5. Копирование листа "аб1" для каждой дороги
    for row_num, row in excel_df.iterrows():
        # Получаем имя для листа
        if row['Номер']:  # пустая строка = False
            sheet_name = row['Номер']
        else:
            # Если номера нет, берем первые 10 символов названия
            sheet_name = str(row['Название']).strip()[:10]

        # Убираем недопустимые символы для имени листа
        invalid_chars = r'[\[\]:\*?/\\]'
        sheet_name = re.sub(invalid_chars, '', sheet_name)

        # Пропускаем, если лист с таким именем уже есть
        if sheet_name not in workbook.sheetnames:
            original_sheet = workbook['аб1']
            new_sheet = workbook.copy_worksheet(original_sheet)
            new_sheet.title = sheet_name

            # Устанавливаем цвет ярлыка по принадлежности дороги
            ownership = row['Принадлежность']
            if ownership in ROAD_COLOR:
                rgb = ROAD_COLOR[ownership]
                r, g, b = map(int, rgb.split(','))
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                new_sheet.sheet_properties.tabColor = hex_color  # цвет ярлыка

        else:
            logger.warning(f"  Лист {sheet_name} уже существует, пропущен")

    # 6. Цветовая заливка строк по принадлежности
    for row_num, row in excel_df.iterrows():
        excel_row = row_num + 2
        ownership = row['Принадлежность']

        if ownership in ROAD_COLOR:
            # строка "255,204,153"
            rgb = ROAD_COLOR[ownership]
            # превращаем в числа (255,204,153)
            r, g, b = map(int, rgb.split(','))
            # конвертируем в HEX "FFCC99"
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            # создаем заливку
            fill = PatternFill(start_color=hex_color,
                               end_color=hex_color, fill_type='solid')

            # Заливаем колонки с D по R
            for cell in worksheet[f'D{excel_row}:R{excel_row}'][0]:
                cell.fill = fill

    # 7. Сохраняем результат
    workbook.save(output_file)
    return output_file